from openpyxl import load_workbook
import json


def import_data1(file_path: str, sheets=[]):
    wb = load_workbook(file_path)
    if not sheets:
        sheets = wb.sheetnames
    for sheet in sheets:
        print(f'--- begin execute [{sheet}]')
        ws = wb[sheet]
        rows = ws.values
        header = next(rows)
        print(f'|-- header: {header}')
        lithest = []
        for row in rows:
            video = dict()
            base = dict()
            base["areaId"] = '321200'
            base["statusId"] = "0"
            for i, h in enumerate(header):
                # ==========video============
                if h == '设备编号':
                    video["id"] = row[i]
                if h == '施工/维保人员':
                    video['responsible'] = row[i]
                if h == '联系电话':
                    video['phone'] = row[i]
                # ==========base============
                if h == '设备IP':
                    base["ip"] = row[i]
                if h == '路口/路段名称':
                    base["address"] = str(row[i])
                if h == '建设点位及方向':
                    base["name"] = str(row[i])
                if h == '设备类型':
                    if '电子警察' in row[i]:
                        base["modelId"] = '0001000300010000'
                    elif '卡' in row[i]:
                        base["modelId"] = '0001000300020000'
                    elif '抓拍' in row[i]:
                        base["modelId"] = '0001000300030000'
                    else:
                        base["modelId"] = '0001000300010000'
                if h == '经度':
                    base['longitude'] = row[i]
                if h == '纬度':
                    base["latitude"] = row[i]
                if h == '所属大队':
                    if '海陵' in row[i]:
                        base["officeId"] = "海陵"
                    elif '高港' in row[i]:
                        base["officeId"] = "高港"
                    elif '高新' in row[i]:
                        base["officeId"] = "高新"
                    elif '快速路' in row[i]:
                        base["officeId"] = "快速路"
                    elif '高一' in row[i]:
                        base["officeId"] = "高一"
                    elif '高二' in row[i]:
                        base["officeId"] = "高二"
                    elif '高三' in row[i]:
                        base["officeId"] = "高三"
                    elif '高四' in row[i]:
                        base["officeId"] = "高四"
                    elif '高五' in row[i]:
                        base["officeId"] = "高五"
                    elif '高六' in row[i]:
                        base["officeId"] = "高六"
                    elif '高七' in row[i]:
                        base["officeId"] = "高七"
                    else:
                        base["officeId"] = "泰州支队"
                if h == '施工单位':
                    if row[i] is None:
                        print("empty maintain")
                    else:
                        if '电信' in str(row[i]):
                            base["maintainId"] = ''
                        if '南京蓝泰' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b98845b44000d'
                        if '江苏尤特斯' in str(row[i]):
                            base["maintainId"] = ''
                        if '南京中盾安防' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b988b3a6b0013'
                        if '泰州诚安' in str(row[i]):
                            base["maintainId"] = ''
                        if '上海宝康' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b9884a655000e'
                        if '南京凌云' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b988294d9000a'
                        if '河北中岗' in str(row[i]):
                            base["maintainId"] = ''
                        if '南京洛普' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b988411c9000c'
                        if '东南智能' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b988602730011'
                        if '长天智远' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b9885687e0010'
                        if '金中天' in str(row[i]):
                            base["maintainId"] = ''
                        if '海阳' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b98917c5e0014'
                        if '兴泰' in str(row[i]):
                            base["maintainId"] = ''
                        if '隆鼎' in str(row[i]):
                            base["maintainId"] = ''
                        if '宏达' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b98832d6a000b'
                        if '鸿信' in str(row[i]):
                            base["maintainId"] = '2c40288b6b78ba5e016b9889f5580012'
            base["deviceId"] = video['id']
            video['base'] = base
            lithest.append(video)
        print(lithest)
    try:
        with open("../res/out/video1.json","w", encoding='utf-8') as fs:
            json.dump(lithest, fs,ensure_ascii=False)
            fs.close()
    except IOError as e:
        print(e)


if __name__ == "__main__":

    import_data1("../res/xlsx/video1.xlsx")