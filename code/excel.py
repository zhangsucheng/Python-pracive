from openpyxl import load_workbook
import json
def main():
    print("读取文件")
    try:
        with open("../res/xlsx/video.xlsx", "rb") as f:
            print(f.read())
    except FileNotFoundError as e:
        print('指定的文件无法打开.')
    except IOError as e:
        print('读写文件时出现错误.')


def main1():
    wb = load_workbook('../res/xlsx/video.xlsx')
    for sheet in wb:
        print(sheet.rows)
        header = next(sheet.rows)
        video = dict()
        for row in sheet.rows:
            for i,h in enumerate(header):
                print(row[i])
                if h == 'IP':
                    video["ip"] = row[i]
                    print(video["ip"])


    wb.close()



def import_data(file_path: str, sheets=[]):
    wb = load_workbook(file_path)
    if not sheets:
        sheets = wb.sheetnames
    for sheet in sheets:
        print(f'--- begin execute [{sheet}]')
        ws = wb[sheet]
        rows = ws.values
        header = next(rows)
        print(f'|-- header: {header}')
        lithest = []
        for row in rows:
            video = dict()
            video["areaId"] = '321200'
            for i, h in enumerate(header):
                if h == 'IP':
                    video["ip"] = row[i]
                if h == '路口/路段名称':
                    video["name"] = str(row[i])
                if h == '建设方向/点位':
                    # excel里如果没有路段路口名称，就用建设点位方向
                    if video["name"] is None:
                        video["name"] = row[i]
                    # excel里只读取简短方向的
                    elif 0 < len(str(row[i])) < 4:
                        video["name"] = str(video["name"]) + row[i]
                    # excel里带特定标识的
                    elif "智能卡口" in str(video["name"]):
                        video["name"] = str(video["name"]) + row[i]
                    else:
                        video["name"] = row[i]
                if h == '设备类型':
                    if row[i] == '电子警察' or row[i] == '普通监控':
                        video["modelId"] = '0001000300010000'
                    if row[i] == '车辆卡口' or row[i] == '测速卡口':
                        video["modelId"] = '0001000300020000'
                if h == '经度':
                    video['longitude'] = row[i]
                if h == '纬度':
                    video["latitude"] = row[i]
                if h == '设备状态':
                    if row[i] == "正常":
                        video["statusId"] = "0"
                    elif row[i] == "维修":
                        video["statusId"] = "1"
                    elif row[i] == "拆除":
                        video["statusId"] = "2"
                    else:
                        video["statusId"] = ""
                if h == '运维单位':
                    if row[i] is None:
                        print("empty maintain")
                    else:
                        if '电信' in str(row[i]):
                            video["maintainId"] = ''
                        if '南京蓝泰' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b98845b44000d'
                        if '江苏尤特斯' in str(row[i]):
                            video["maintainId"] = ''
                        if '南京中盾安防' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b988b3a6b0013'
                        if '泰州诚安' in str(row[i]):
                            video["maintainId"] = ''
                        if '上海宝康' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b9884a655000e'
                        if '南京凌云' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b988294d9000a'
                        if '河北中岗' in str(row[i]):
                            video["maintainId"] = ''
                        if '南京洛普' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b988411c9000c'
                        if '东南智能' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b988602730011'
                        if '长天智远' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b9885687e0010'
                        if '金中天' in str(row[i]):
                            video["maintainId"] = ''
                        if '海阳' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b98917c5e0014'
                        if '兴泰' in str(row[i]):
                            video["maintainId"] = ''
                        if '隆鼎' in str(row[i]):
                            video["maintainId"] = ''
                        if '宏达' in str(row[i]):
                            video["maintainId"] = '2c40288b6b78ba5e016b98832d6a000b'
            lithest.append(video)
        print(lithest)
    try:
        with open("../res/out/video.json","w", encoding='utf-8') as fs:
            json.dump(lithest, fs,ensure_ascii=False)
            fs.close()
    except IOError as e:
        print(e)



if __name__ == "__main__":
    import_data("../res/xlsx/video.xlsx")
